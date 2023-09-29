from xlsxwriter import *

from database import *

import xlrd
import datetime

import time
from utils import get_coordinates
import openpyxl

#добавляем все данные в базу


if __name__=="__main__":
    #добавляем все города в базу
    wookbook = openpyxl.load_workbook("Лист Microsoft Excel.xlsx")

    worksheet = wookbook.active

    for row in worksheet.iter_rows(min_row=1, max_col=7, max_row=worksheet.max_row):    # row=worksheet.row
        name, type, kladr_code, index, ghin,uno, is_capital = row

        name=name.value
        type = type.value
        kladr_code = kladr_code.value
        
        settlement=Settlement(Name=name,
                            Latitude=0,
                            Longitude=0,
                            KladrIndex=kladr_code,
                            Type=type)
        Session.add(settlement)
    Session.commit()
    punkts=Session.query(Settlement).all()

    # записываем координаты в бд( происходит долго,поэтому в случае выключения пк, все данные запишутся в бд )
    try:

        for punkt in punkts:
                if int(punkt.Latitude)==0 and int(punkt.Longitude)==0:
                    name=punkt.Name
                    type=punkt.Type
                    if type=="д":
                         type="деревня"
                    if type=="х":
                         type="хутор"
                    if type=="г-к":
                         type="городок"
                    coordinates=get_coordinates(type+" "+name)
                    latitude=coordinates["latitude"]
                    longitude=coordinates["longitude"]
                    Session.query(Settlement).filter(Settlement.Name==name).update({"Longitude":longitude,"Latitude":latitude})
                    time.sleep(0.5)
    except Exception as err:
        print (err)
        time.sleep(10)

    finally:
        Session.commit()
